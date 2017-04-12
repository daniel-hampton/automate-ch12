#! python3
"""cell_inverter.py transposes the array of cells in a spreadsheet. Saves new copy of spreadsheet."""

import sys
import os
import openpyxl


# get arguments from command line
if len(sys.argv) > 1:
    try:
        filepath = os.path.abspath(sys.argv[1])
    except Exception as err:
        print('Something is wrong with file provide:  {}'.format(err))
        sys.exit(99)
else:
    print('Must enter filename for program arguments.')
    sys.exit(98)

wb = openpyxl.load_workbook(filepath)  # load source spreadsheet
ws = wb.active

sheetData = tuple(ws.values)  # extract sheet data

wb2 = openpyxl.Workbook()  # create destination workbook in memory
ws2 = wb2.active

# Loop through cells in sheetData[x][y] writing to ws2 cells[y][x]
for rowNum in range(len(sheetData)):
    for colNum in range(len(sheetData[rowNum])):
        ws2.cell(row=colNum + 1, column=rowNum + 1).value = sheetData[rowNum][colNum]

# save workbook as separate file
filename = os.path.basename(filepath).rsplit('.', 1)
wb2.save(''.join([filename[0], '_transposed.', filename[1]]))
