#! /usr/bin/python3
# multitables.py - creates a multiplication table from inputs from the commandline

import sys
import openpyxl

# read arguments from command line

if len(sys.argv) > 1:
	try:
		num = int(sys.argv[1])
	except ValueError:
		print('That was not a valid integer')
elif len(sys.argv) == 1:
	print('Needs a number N after file')

wb = openpyxl.Workbook()

sheet = wb.get_sheet_by_name('Sheet')

sheet = wb.active

numList = [x for x in range(1, num + 1)]

for x in range(1, num + 1):
	sheet['A{}'.format(x + 1)].value = x

for x in range(1, num + 1):
	sheet.cell(row=1, column=(x + 1)).value = x

# TODO: Multiply numbers
for colNum in range(2, sheet.max_column + 1):
	for rowNum in range(2, sheet.max_row + 1):
		product = sheet.cell(row=rowNum, column=1).value * sheet.cell(row=1, column=colNum).value
		sheet.cell(row=rowNum, column=colNum).value = product

wb.save('tables.xlsx')
