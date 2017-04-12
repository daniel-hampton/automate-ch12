#! /usr/bin/python3
# blankRowIns.py - Take arguments from the command line, N M filename.xlsx
# where N is the row to insert M blank rows in filename.xlsx

import sys
import openpyxl


if len(sys.argv) > 1:
    try:
        N = int(sys.argv[1])
        M = int(sys.argv[2])
    except ValueError:
        print('Arguments must be in format: N M filename.xlsx where N and M\
         are integers')
        raise
    try:
        filename = ' '.join(sys.argv[3:])
    except Exception as err:
        print('An error occured with filename: {}'.format(err))
        raise

# create workbook
wb = openpyxl.load_workbook(filename)
sheet = wb.active

wb2 = openpyxl.Workbook()
sheet_wb2 = wb2.get_sheet_by_name('Sheet')

# for rows 1 through N copy the data directly to sheet in new workbook object
for rowNum in range(1, N + 1):
    for colNum in range(1, sheet.max_column + 1):
        sheet_wb2.cell(row=rowNum, column=colNum).value \
        = sheet.cell(row=rowNum,column=colNum).value

# for rows N+1 through end of file, copy to wb as rowNum + M. add M blank rows                                                               
for rowNum in range(N + 1, sheet.max_row + 1):
    for colNum in range(1, sheet.max_column + 1):
        sheet_wb2.cell(row=rowNum + M, column=colNum).value \
        = sheet.cell(row=rowNum,column=colNum).value

wb2.save('{}_copy.xlsx'.format(filename.rsplit('.', 1)[0]))
